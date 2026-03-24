from openpyxl import load_workbook, Workbook
from pathlib import Path

input_file = Path("output.xlsx")

output_dir = input_file.parent / "extracted_sheets"
output_dir.mkdir(exist_ok=True)

wb = load_workbook(input_file, read_only=True)

def safe_name(name):
    invalid = '<>:"/\\|?*'
    return "".join("_" if c in invalid else c for c in name).strip() or "Sheet"

for sheet in wb.worksheets:
    new_wb = Workbook(write_only=True)
    new_ws = new_wb.create_sheet(title=sheet.title[:31])

    for row in sheet.iter_rows(values_only=True):
        new_ws.append(row)

    file_name = safe_name(sheet.title) + ".xlsx"
    output_path = output_dir / file_name

    new_wb.save(output_path)

    print(f"Saved: {output_path}")

wb.close()
