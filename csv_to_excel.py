import csv
from pathlib import Path

from openpyxl import Workbook


CSV_FILE = Path("csv.csv")
EXCEL_FILE = Path("output.xlsx")
DELIMITER = ","
ENCODING = "utf-8"
EXCEL_MAX_ROWS = 1_048_576
PROGRESS_EVERY_ROWS = 200_000


def csv_to_excel_streaming(csv_path: Path, excel_path: Path) -> tuple[int, int]:
    """Convert CSV to XLSX with low memory usage and automatic sheet splitting."""
    workbook = Workbook(write_only=True)
    sheet_number = 1
    worksheet = workbook.create_sheet(title=f"Sheet{sheet_number}")
    rows_in_sheet = 0
    total_rows = 0

    with csv_path.open("r", encoding=ENCODING, errors="replace", newline="") as source:
        reader = csv.reader(source, delimiter=DELIMITER)

        for row in reader:
            if rows_in_sheet >= EXCEL_MAX_ROWS:
                sheet_number += 1
                worksheet = workbook.create_sheet(title=f"Sheet{sheet_number}")
                rows_in_sheet = 0

            worksheet.append(row)
            rows_in_sheet += 1
            total_rows += 1

            if total_rows % PROGRESS_EVERY_ROWS == 0:
                print(f"Processed {total_rows:,} rows...")

    workbook.save(excel_path)
    return total_rows, sheet_number


def main() -> None:
    if not CSV_FILE.exists():
        raise FileNotFoundError(f"CSV file not found: {CSV_FILE.resolve()}")

    rows, sheets = csv_to_excel_streaming(CSV_FILE, EXCEL_FILE)
    print(f"Conversion completed: {rows:,} rows written across {sheets} sheet(s).")
    print(f"Output file: {EXCEL_FILE.resolve()}")


if __name__ == "__main__":
    main()
