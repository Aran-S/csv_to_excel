from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel


SOURCE_DIR = Path("split_sheets")
OUTPUT_DIR = Path("extracted_sheets")
OUTPUT_FILE = OUTPUT_DIR / "incorporation_jan24_dec24.xlsx"

START_DATE = date(2024, 1, 1)
END_DATE = date(2024, 12, 31)

KEYWORDS = ("incorporated", "incor
"
"poration date")


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None

    if not isinstance(value, str):
        return None

    text = value.strip()
    if not text:
        return None

    formats = [
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%d %b %Y",
        "%d %B %Y",
        "%b %d %Y",
        "%B %d %Y",
        "%b-%y",  # Jan-24
        "%b %y",  # Jan 24
        "%b-%Y",
        "%b %Y",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def row_contains_keyword(row):
    row_text = " ".join(str(cell).lower() for cell in row if cell is not None)
    return any(keyword in row_text for keyword in KEYWORDS)


def row_has_2024_date(row):
    for cell in row:
        parsed = parse_date(cell)
        if parsed and START_DATE <= parsed <= END_DATE:
            return True
    return False


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)

    files = sorted(SOURCE_DIR.glob("*.xlsx"))
    if not files:
        print(f"No Excel files found in: {SOURCE_DIR}")
        return

    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    total_matches = 0

    for file_path in files:
        in_wb = load_workbook(file_path, data_only=True, read_only=True)
        in_ws = in_wb.active

        out_ws = out_wb.create_sheet(title=file_path.stem[:31])

        rows = in_ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is not None:
            out_ws.append(header)

        file_matches = 0

        for row in rows:
            if row_contains_keyword(row) and row_has_2024_date(row):
                out_ws.append(row)
                file_matches += 1

        total_matches += file_matches
        in_wb.close()
        print(f"{file_path.name}: {file_matches} matching rows")

    out_wb.save(OUTPUT_FILE)
    print(f"Saved: {OUTPUT_FILE}")
    print(f"Total matching rows: {total_matches}")


if __name__ == "__main__":
    main()
